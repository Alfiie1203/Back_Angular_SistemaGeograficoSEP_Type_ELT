from cities_light.models import Country, Region, SubRegion
from apps.company.models.actor_type import ActorType
from apps.company.models.company import Company
from apps.company.models.commodity import Commodity
from apps.supplybase.models.supplybase import SupplyBaseDependency

from apps.company.models.company_group import CompanyGroup
from apps.formulario.models.formulario import Period

from apps.user.models import User
from django.http import JsonResponse
from django.views.generic import View
from django.http import HttpResponse
from django.utils import translation


from ..models.traceability import Traceability, TraceabilityFile

from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from datetime import datetime

__location__ = os.path.realpath(os.path.join(
    os.getcwd(), os.path.dirname(__file__)))


# =================================================================================
# =========               Export Template FILE AS  VIEW         ==============
# =================================================================================

class XlsxExportView(View):
    url_name = 'xlsx-export'

    def get(self, request, *args, **kwargs):
        id_empresa_reportada = self.request['id_empresa_reporta']
        name_translate = 'name_'+ str(translation.get_language())
        # Load existing workbook.
        file_xlsx = open(os.path.join(
            __location__, 'plantilla_traceabilty_final.xlsx'))
        workbook = openpyxl.load_workbook(filename=file_xlsx.name)

        # Write a Commodity that belongs company of current user company or admin request
        if self.request.user.role.name=='COLABORADOR':
            reported_company = Company.objects.get(id=id_empresa_reportada)
        else:
            reported_company = request.user.company

        commodity = reported_company.commodity
        sheet_fill = workbook['plantilla']
        sheet_fill.cell(2, 1).value = getattr(commodity, name_translate)

        # Write Actortypes asociated
        worksheet = workbook['data']
        actortypes = SupplyBaseDependency.objects.get(actor_type = reported_company.actor_type).actor_type_dependency.all()
        lista_actors = list(actortypes.values_list(name_translate, flat=True))
        row = 2
        for actor in lista_actors:
            worksheet.cell(row, 1).value = actor
            row += 1

        # Write Period from database-model
        worksheet = workbook['data']
        periods = [0,1,2]
        row = 2
        for period in periods:
            worksheet.cell(row, 2).value = period
            row += 1

        # Set up the Http response.
        response = HttpResponse(content=save_virtual_workbook(workbook))
        response['Content-Disposition'] = f'attachment; filename={getattr(commodity, name_translate)}_template_traceability.xlsx'

        return response

@permission_classes([IsAuthenticated])
class XlsxExportOthersView(View):
    url_name = 'xlsx-export-others'

    def get(self, request, *args, **kwargs):
        name_translate = 'name_'+ str(translation.get_language())
        id_empresa_reportada = self.request['id_empresa_reporta']
        # Load existing workbook.
        file_xlsx = open(os.path.join(__location__, 'plantilla_otros.xlsx'))
        workbook = openpyxl.load_workbook(filename=file_xlsx.name)
        
        # Write a Commodity that belongs company of current user company or admin request
        if self.request.user.role.name=='COLABORADOR':
            reported_company = Company.objects.get(id=id_empresa_reportada)
        else:
            reported_company = request.user.company

        commodity = reported_company.commodity
        sheet_fill = workbook['plantilla']
        sheet_fill.cell(2, 1).value = getattr(commodity, name_translate)

        # Write Actortypes asociated
        worksheet = workbook['data']
        actortypes = SupplyBaseDependency.objects.get(actor_type = reported_company.actor_type).actor_type_dependency.all()
        lista_actors = list(actortypes.values_list(name_translate, flat=True))

        row = 2
        for actor in lista_actors:
            worksheet.cell(row, 1).value = actor
            row += 1

        # Write Period from database-model
        worksheet = workbook['data']
        periods = [0,1,2]
        row = 2
        for period in periods:
            worksheet.cell(row, 2).value = period
            row += 1

        # Set up the Http response.
        response = HttpResponse(content=save_virtual_workbook(workbook))
        response['Content-Disposition'] = f'attachment; filename={getattr(commodity, name_translate)}_template_traceability.xlsx'

        return response

# =================================================================================
# =========               Export Template FILE AS API-REST         ============
# =================================================================================

@permission_classes([IsAuthenticated])
class XlsxExportApi(APIView):

    def get(self, request, *args, **kwargs):
        name_translate = 'name_'+ str(self.request.headers.get('Content-Language'))
        id_empresa_reportada = self.request.query_params.get('id_empresa_reporta')
        if self.request.user.role.name=='COLABORADOR':
            reported_company = Company.objects.get(id=id_empresa_reportada)
        else:
            reported_company = request.user.company
        # Load existing workbook.
        file_xlsx = open(os.path.join(
            __location__, 'plantilla_traceabilty_final.xlsx'))
        workbook = openpyxl.load_workbook(filename=file_xlsx.name)

        # Write a Commodity that belongs company of current user
        try:
            commodity = reported_company.commodity
        except:
            return Response(
                {

                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "not_found",
                            "detail": "Company don't has commodity to generate Traceability Template commodity",
                            "attr": "Traceability"
                        }
                    ]
                },
                404
            )
        sheet_fill = workbook['plantilla']
        sheet_fill.cell(2, 1).value = getattr(commodity, name_translate)

        # Write Actortypes asociated
        worksheet = workbook['data']
        actortypes = SupplyBaseDependency.objects.get(actor_type = reported_company.actor_type).actor_type_dependency.all()
        lista_actors = list(actortypes.values_list(name_translate, flat=True))
        row = 2
        for actor in lista_actors:
            worksheet.cell(row, 1).value = actor
            row += 1
        
        # Write Period from database-model
        worksheet = workbook['data']
        periods = [0,1,2]
        row = 2
        for period in periods:
            worksheet.cell(row, 2).value = period
            row += 1

        # Set up the Http response.
        response = HttpResponse(save_virtual_workbook(
            workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template_traceability.xlsx'
        return response

@permission_classes([IsAuthenticated])
class XlsxExportOthersApi(APIView):

    def get(self, request, *args, **kwargs):
        name_translate = 'name_'+ str(self.request.headers.get('Content-Language'))
        id_empresa_reportada = self.request.query_params.get('id_empresa_reporta', None)
        if self.request.user.role.name=='COLABORADOR':
            reported_company = Company.objects.get(id=id_empresa_reportada)
        else:
            reported_company = request.user.company
        # Load existing workbook.
        file_xlsx = open(os.path.join(__location__, 'plantilla_otros.xlsx'))
        workbook = openpyxl.load_workbook(filename=file_xlsx.name)

        # Write a Commodity that belongs company of current user
        try:
            commodity = reported_company.commodity
        except:
            return Response(
                {

                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "not_found",
                            "detail": "User does not have a company to generate Traceability Other Actors Template commodity",
                            "attr": "Traceability"
                        }
                    ]
                },
                404
            )

        sheet_fill = workbook['plantilla']
        sheet_fill.cell(2, 1).value = getattr(commodity, name_translate)

        # Write Actortypes asociated
        worksheet = workbook['data']
        actortypes = SupplyBaseDependency.objects.get(actor_type = reported_company.actor_type).actor_type_dependency.all()
        lista_actors = list(actortypes.values_list(name_translate, flat=True))
        row = 2
        for actor in lista_actors:
            worksheet.cell(row, 1).value = actor
            row += 1

        # Write Period from database-model
        worksheet = workbook['data']
        periods = [0,1,2]
        row = 2
        for period in periods:
            worksheet.cell(row, 2).value = period
            row += 1

        # Set up the Http response.
        response = HttpResponse(save_virtual_workbook(
            workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template_traceability.xlsx'
        return response

# =================================================================================
# =========           EXPORT  XLSX  TRACEABILITY FILE     ============
# =================================================================================

@permission_classes([IsAuthenticated])
class XlsxExportTraceabilityView(View):
    url_name = 'xlsx-traceability-export'

    def get(self, request, *args, **kwargs):
        queryset = Traceability.objects.all()
        company_name = self.request.GET.get('company_name')
        created_at_gte = self.request.GET.get('created_at_gte')
        created_at_lte = self.request.GET.get('created_at_lte')
        if created_at_gte:
            date_reported = datetime.strptime(created_at_gte, '%Y-%m-%d %H:%M')
            queryset = queryset.filter(
                date_reported__gte=date_reported)

        if created_at_lte:
            date_reported = datetime.strptime(created_at_lte, '%Y-%m-%d %H:%M')
            queryset = queryset.filter(
                date_reported__lte=date_reported)
        
        if company_name:
            queryset = queryset.filter(
                supplier_name=company_name)
        
        wb = Workbook()
        worksheet = wb.active
        worksheet.title = "Traceability"
        columns = [
            'Fecha',
            'Commodity',
            'Actor type',
            'Group Name',
            'Supplier Name',
            'Supplier Tax Number',
            'Area ha (farm)',
            'Area ha (plantation)',
            'Purchased volume (tipo %)',
            'Certification',
            'Latitude',
            'Longitude',
            'Country',
            'Province',
            'District',
            'Year',
            'Period'

        ]
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for _, trace in enumerate(queryset, 1):
            row_num += 1
            row = [
                trace.date_reported.replace(tzinfo=None),
                trace.commodity.name,
                trace.actor_type.name,
                trace.company_group.name if trace.company_group else None,
                trace.supplier_name,
                trace.supplier_tax_number,
                trace.supplier_capacity,
                trace.supplier_production,
                trace.purchased_volume,
                trace.certification,
                trace.latitude,
                trace.longitude,
                trace.country.name,
                trace.region.name,
                trace.city.name,
                trace.year,
                trace.period,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        response = HttpResponse(content=save_virtual_workbook(wb))
        response['Content-Disposition'] = 'attachment; filename=traceability.xlsx'

        return response


# =============================================================================
#                      Read File for Create Traceabiliyty 
# =============================================================================

@permission_classes([IsAuthenticated])
class TraceabilityCreate(APIView):

    def get(self, request):
        pass

    def post(self, request):
        id_empresa_reportada = self.request.query_params.get('id_empresa_reporta', None)
        try:
            if self.request.user.role.name=='COLABORADOR':
                reported_company = Company.objects.get(id=id_empresa_reportada)
            else:
                reported_company = request.user.company

        except:
            return Response( {
                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "wrong_id_empresa_reporta",
                            "detail": "The request must be from an client or contain a valid id_empresa_reporta ",
                            "attr": "company"
                        }
                    ]
                },
                403)
        file = request.FILES['file_traceability']
        format_file = file.name.split('.')[-1]
        name_translate = 'name_'+ str(self.request.headers.get('Content-Language'))

        if (format_file in ['xls', 'xlsx']):
            try:
                file.name = str(self.request.user.company.identifier_global_company)+('.')+format_file
                file.save()
            except:
                pass
            TraceabilityFile.objects.create(
                upload_by=self.request.user,
                reporting_company=self.request.user.company,
                file_traceability=file
            )
        else:
            return Response(
                {

                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "wrong_format",
                            "detail": "The uploaded file must be contain an .XLS or .XLSX extension. Template data must be in first sheet1",
                            "attr": "traceability"
                        }
                    ]
                },
                403
            )

        workbook = load_workbook(filename=file,  read_only=True)
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]
        load_result = []
        processed_register = 0
        omitted_register = 0
        for i, row in enumerate(worksheet.iter_rows()):
            if not i == 0 and row[0].value != None:
                lookup = "%s__contains" % name_translate
                
                dictionary = clean_data_row(row, lookup, name_translate)

                if dictionary['validated'] == True:
                    suppliercompany = dictionary['company']
                    was_company_in_database = dictionary['was_company_in_database']
                    if Traceability.objects.filter(year=row[14].value, period=row[15].value, reported_company = reported_company, supplier_company=suppliercompany ).exists():
                        traceability_update = Traceability.objects.filter(year=row[14].value, period=row[15].value, reported_company = reported_company, supplier_company=suppliercompany)[0]
                        traceability_update.reported_user=self.request.user
                        traceability_update.company_group=suppliercompany.company_group
                        traceability_update.supplier_name=suppliercompany.name
                        traceability_update.supplier_tax_number=row[4].value
                        traceability_update.supplier_capacity=row[5].value
                        traceability_update.supplier_production=row[6].value
                        traceability_update.purchased_volume=row[7].value
                        traceability_update.certification=row[8].value
                        traceability_update.latitude=float(row[9].value)
                        traceability_update.longitude=float(row[10].value)
                        traceability_update.country= dictionary['country_traceability']
                        traceability_update.region= dictionary['region_traceability']
                        traceability_update.city= dictionary['city_traceability']
                        traceability_update.year=row[14].value
                        traceability_update.period=row[15].value
                        traceability_update.save()
                        status_register = f'Traceability record for this year and period updated'
                        processed_register += 1
                        status = True
                    else:
                        Traceability.objects.create(
                            reported_user=self.request.user,
                            reported_company = reported_company,
                            supplier_company=suppliercompany,
                            commodity=suppliercompany.commodity,
                            actor_type=suppliercompany.actor_type,
                            company_group=suppliercompany.company_group,
                            supplier_name=suppliercompany.name,
                            supplier_tax_number=row[4].value,
                            supplier_capacity=row[5].value,
                            supplier_production=row[6].value,
                            purchased_volume=row[7].value,
                            certification=row[8].value,
                            latitude=float(row[9].value),
                            longitude=float(row[10].value),
                            country= dictionary['country_traceability'],
                            region= dictionary['region_traceability'],
                            city= dictionary['city_traceability'],
                            year=row[14].value,
                            period=row[15].value,
                        )
                        processed_register += 1
                        status_register = f'Register created'
                        status = True
                        if (suppliercompany.company_group == None) and (row[2].value != None):
                            status_register = status_register + \
                                f' , but company group are  not created or updated on data base '
                else:
                    error_text = dictionary['error']
                    status_register = f'Register ommited {error_text}'
                    omitted_register += 1
                    status = False
                    was_company_in_database = "Register with error"

                load_result.append({"index": i,
                                    "status": status,
                                    "note": str(status_register),
                                    "was_company_in_database": was_company_in_database
                                     })

        return JsonResponse({"load_result": load_result,
                            "Processed registers": processed_register,
                            "Registers with errors": omitted_register,
                            "Total Registers ": processed_register + omitted_register}, status=200)


def clean_data_row(row, lookup, language_code):
    try:
        commodity = Commodity.objects.filter(**{ lookup: row[0].value.strip()})[0]
        actor_type = ActorType.objects.filter(commodity=commodity).filter(**{ lookup: row[1].value.strip()})[0]
        nit = str(row[4].value).strip()  # clear spaces
        latitude = float(row[9].value)
        longitude = float(row[10].value)
        country_name = row[11].value.strip()
        region_name = row[12].value.strip()
        city_name = row[13].value.strip()
        company_group = CompanyGroup.objects.get(name=row[2].value.strip()) if CompanyGroup.objects.filter(name=row[2].value).exists() else None
        country_suplier = Country.objects.filter(name__unaccent__icontains=country_name)[0]
        region_suplier = Region.objects.filter(country=country_suplier, name__unaccent__icontains=region_name)[0]
        city_suplier = SubRegion.objects.filter(region=region_suplier, name__unaccent__icontains=city_name)[0]
        
        if Company.objects.filter(nit=nit).exists():

            if Company.objects.filter(nit=nit, commodity=commodity, actor_type=actor_type).exists():
                suppliercompany = Company.objects.get(
                    nit=nit, commodity=commodity, actor_type=actor_type)
                was_company_in_database =  "Company registered"

            else: # create a suplier company
                suppliercompany = Company.objects.create(
                    name=row[3].value.strip(),
                    country=country_suplier,
                    region=region_suplier,
                    city=city_suplier,
                    latitude=latitude,
                    longitude=longitude,
                    nit=nit,
                    commodity=commodity,
                    actor_type=actor_type,
                    company_group=company_group,
                    company_profile='SC'
                )
                was_company_in_database = "New Company"


        else: # create a suplier company
            suppliercompany = Company.objects.create(
                name=row[3].value.strip(),
                country=country_suplier,
                region=region_suplier,
                city=city_suplier,
                latitude=latitude,
                longitude=longitude,
                nit=nit,
                commodity=commodity,
                actor_type=actor_type,
                company_group=company_group,
                company_profile='SC'
            )
            was_company_in_database = "New Company"

        return ({'validated': True,
                'company': suppliercompany,
                'was_company_in_database': was_company_in_database,
                'country_traceability': country_suplier,
                'region_traceability' : region_suplier,
                'city_traceability' : city_suplier
         })

    except Exception as error:
        return ({'validated': False, 'error': 'Not valid row'})


@permission_classes([IsAuthenticated])
class TraceabilityOtherActorCreate(APIView):
    def get(self, request):
        pass

    def post(self, request):
        id_empresa_reportada = self.request.query_params.get('id_empresa_reporta', None)
        try:
            if self.request.user.role.name=='COLABORADOR':
                reported_company = Company.objects.get(id=id_empresa_reportada)
            else:
                reported_company = request.user.company

        except:
            return Response( {
                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "wrong_id_empresa_reporta",
                            "detail": "The request must be from an client or contain a valid id_empresa_reporta ",
                            "attr": "company"
                        }
                    ]
                },
                403)
        file = request.FILES['file_traceability']
        format_file = file.name.split('.')[-1]
        name_translate = 'name_'+ str(self.request.headers.get('Content-Language'))

        if (format_file in ['xls', 'xlsx']):
            try:
                file.name = str(self.request.user.company.identifier_global_company)+('.')+format_file
                file.save()
            except:
                pass
            TraceabilityFile.objects.create(
                upload_by=self.request.user,
                reporting_company=self.request.user.company,
                file_traceability=file
            )
        else:
            return Response(
                {
                    "type": "validation_error",
                    "errors": [
                        {
                            "code": "wrong_format",
                            "detail": "The uploaded file must be contain an .XLS or .XLSX extension.",
                            "attr": "traceability"
                        }
                    ]
                },
                403
            )
        workbook = load_workbook(filename=file,  read_only=True)
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]
        load_result = []
        processed_register = 0
        omitted_register = 0
        for i, row in enumerate(worksheet.iter_rows()):
            if not i == 0 and row[0].value != None:
                lookup = "%s__contains" % name_translate
                
                dictionary = clean_data_row_other_actors(row, lookup, name_translate)

                if dictionary['validated'] == True:
                    suppliercompany = dictionary['company']
                    was_company_in_database = dictionary['was_company_in_database']
                    if Traceability.objects.filter(year=row[11].value, period=row[12].value, reported_company=reported_company, supplier_company=suppliercompany).exists():
                        traceability_update = Traceability.objects.filter(year=row[11].value, period=row[12].value, reported_company=reported_company, supplier_company=suppliercompany)[0]
                        traceability_update.reported_user=self.request.user
                        traceability_update.supplier_company=suppliercompany
                        traceability_update.commodity=suppliercompany.commodity
                        traceability_update.actor_type=suppliercompany.actor_type
                        traceability_update.company_group=suppliercompany.company_group
                        traceability_update.supplier_name=suppliercompany.name
                        traceability_update.supplier_tax_number=row[4].value
                        traceability_update.purchased_volume=row[5].value
                        traceability_update.latitude=float(row[6].value)
                        traceability_update.longitude=float(row[7].value)
                        traceability_update.country= dictionary['country_traceability']
                        traceability_update.region= dictionary['region_traceability']
                        traceability_update.city= dictionary['city_traceability']
                        traceability_update.year=row[11].value
                        traceability_update.period=row[12].value
                        traceability_update.save()
                        status_register = f'Traceability record for this year and period updated'
                        processed_register += 1
                        status = True
                    else:
                        Traceability.objects.create(
                            reported_user=self.request.user,
                            reported_company = reported_company,
                            supplier_company=suppliercompany,
                            commodity=suppliercompany.commodity,
                            actor_type=suppliercompany.actor_type,
                            company_group=suppliercompany.company_group,
                            supplier_name=suppliercompany.name,
                            supplier_tax_number=row[4].value,
                            purchased_volume=row[5].value,
                            latitude=float(row[6].value),
                            longitude=float(row[7].value),
                            country= dictionary['country_traceability'],
                            region= dictionary['region_traceability'],
                            city= dictionary['city_traceability'],
                            year=row[11].value,
                            period=row[12].value,
                        )
                        processed_register += 1
                        status_register = f' Register created'
                        status = True
                        if (suppliercompany.company_group == None) and (row[2].value != None):
                            status_register = status_register + \
                                f', but company group are  not created or updated on data base '
                else:
                    error_text = dictionary['error']
                    status_register = f' Register ommited {error_text}'
                    omitted_register += 1
                    status = False
                    was_company_in_database = "Register with error"

                load_result.append({"index": i,
                                    "status": status,
                                    "note": str(status_register),
                                    "was_company_in_database": was_company_in_database
                                     })

        return JsonResponse({"load_result": load_result,
                            "Processed registers": processed_register,
                            "Registers with errors": omitted_register,
                            "Total Registers ": processed_register + omitted_register}, status=200)


def clean_data_row_other_actors(row, lookup, language_code):
    try:
        commodity = Commodity.objects.filter(**{ lookup: row[0].value.strip()})[0]
        actor_type = ActorType.objects.filter(commodity=commodity).filter(**{ lookup: row[1].value.strip()})[0]
        nit = str(row[4].value).strip()  # clear spaces
        latitude = float(row[6].value)
        longitude = float(row[7].value)

        country_name = row[8].value.strip()
        region_name = row[9].value.strip()
        city_name = row[10].value.strip()
        company_group = CompanyGroup.objects.get(name=row[2].value.strip()) if CompanyGroup.objects.filter(name=row[2].value).exists() else None
        country_suplier = Country.objects.filter(name__unaccent__icontains=country_name)[0]
        region_suplier = Region.objects.filter(country=country_suplier, name__unaccent__icontains=region_name)[0]
        city_suplier = SubRegion.objects.filter(region=region_suplier, name__unaccent__icontains=city_name)[0]

        if Company.objects.filter(nit=nit).exists():
            if Company.objects.filter(nit=nit, commodity=commodity, actor_type=actor_type).exists():
                suppliercompany = Company.objects.get(
                    nit=nit, commodity=commodity, actor_type=actor_type)
                was_company_in_database =  "Company registered"
            
            else:# create a suplier company
                suppliercompany = Company.objects.create(
                    name=row[3].value.strip(),
                    country=country_suplier,
                    region=region_suplier,
                    city=city_suplier,
                    latitude=latitude,
                    longitude=longitude,
                    nit=nit,
                    commodity=commodity,
                    actor_type=actor_type,
                    company_group=company_group,
                    company_profile='SO'
                )
                was_company_in_database = "New Company"

        else:
            suppliercompany = Company.objects.create(
                name=row[3].value.strip(),
                country=country_suplier,
                region=region_suplier,
                city=city_suplier,
                latitude=latitude,
                longitude=longitude,
                nit=nit,
                commodity=commodity,
                actor_type=actor_type,
                company_group=company_group,
                company_profile='SO'
            )
            was_company_in_database = "New Company"

        return ({'validated': True,
                'company': suppliercompany,
                'was_company_in_database': was_company_in_database,
                'country_traceability': country_suplier,
                'region_traceability' : region_suplier,
                'city_traceability' : city_suplier
                })

    except Exception as error:
        return ({'validated': False, 'error': 'Not valid row'})

# =============================================================


